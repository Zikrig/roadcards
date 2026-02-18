from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from bot.keyboards import (
    get_admin_main_menu,
    get_report_type_kb,
    get_confirm_format_kb,
    get_documents_kb,
)
from bot.utils import update_last_update_time
from database.db import add_to_whitelist, async_session
from database.models import Transaction, TransactionType
import pandas as pd
import io
import os
import openpyxl
from datetime import datetime
from sqlalchemy import select, and_
from config import ADMIN_IDS

router = Router()

class AdminState(StatesGroup):
    waiting_for_expense_file = State()
    waiting_for_payment_file = State()
    waiting_for_export_start_date = State()
    waiting_for_export_end_date = State()
    waiting_for_broadcast_msg = State()
    confirm_format_expense = State()
    confirm_format_payment = State()
    waiting_for_link_card = State()
    waiting_for_document_choice = State()

def is_admin(user_id: int):
    return user_id in ADMIN_IDS

@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer("Админ-панель:", reply_markup=get_admin_main_menu())

@router.callback_query(F.data == "admin_upload")
async def process_upload_report(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return
    await callback.message.edit_text("Выберите тип отчета:", reply_markup=get_report_type_kb())
    await callback.answer()

@router.callback_query(F.data == "admin_docs")
async def list_documents(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return

    async with async_session() as session:
        result = await session.execute(
            select(Transaction.document)
            .where(Transaction.document.is_not(None))
            .distinct()
            .order_by(Transaction.document)
        )
        docs = [d for (d,) in result.all() if d]

    if not docs:
        await callback.message.edit_text("Документы (дампы) не найдены.")
        await callback.answer()
        return

    await state.update_data(documents=docs)
    await callback.message.edit_text(
        "Выберите документ, который нужно отозвать:",
        reply_markup=get_documents_kb(docs),
    )
    await state.set_state(AdminState.waiting_for_document_choice)
    await callback.answer()

@router.callback_query(AdminState.waiting_for_document_choice, F.data.startswith("admin_doc_"))
async def revoke_document(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return

    idx_str = callback.data.split("_")[-1]
    try:
        idx = int(idx_str)
    except ValueError:
        await callback.answer("Неверный формат выбора.")
        return

    data = await state.get_data()
    docs = data.get("documents", [])
    if idx < 0 or idx >= len(docs):
        await callback.answer("Документ не найден.")
        return

    document_label = docs[idx]

    async with async_session() as session:
        result = await session.execute(
            select(Transaction).where(Transaction.document == document_label)
        )
        txs = result.scalars().all()
        deleted_count = len(txs)
        for t in txs:
            await session.delete(t)
        await session.commit()

    await callback.message.edit_text(
        f"Документ «{document_label}» отозван. "
        f"Удалено {deleted_count} записей из базы данных."
    )
    await state.clear()
    await callback.answer()

@router.callback_query(F.data == "report_expense")
async def report_expense_cb(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Пожалуйста, загрузите Excel файл с ТРАТАМИ клиентов.")
    await state.set_state(AdminState.waiting_for_expense_file)
    await callback.answer()

@router.callback_query(F.data == "report_payment")
async def report_payment_cb(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Пожалуйста, загрузите Excel файл с ОПЛАТАМИ клиентов.")
    await state.set_state(AdminState.waiting_for_payment_file)
    await callback.answer()

async def validate_format(file_content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # B2 is row 2, col 2; A3 is row 3, col 1; B3 is row 3, col 2
    b2_val = ws.cell(row=2, column=2).value
    a3_val = ws.cell(row=3, column=1).value
    b3_val = ws.cell(row=3, column=2).value
    
    is_valid = (b2_val is None or str(b2_val).strip() == "") and \
               (a3_val is None or str(a3_val).strip() == "") and \
               (b3_val is not None and str(b3_val).strip() != "")
               
    return is_valid

async def process_excel_expense(file_content: bytes, document: str):
    # Format: Фирма, карта, дата, адрес, наименование, количество, цена, стоимость
    df = pd.read_excel(io.BytesIO(file_content), skiprows=2, usecols="B:I")
    df.columns = ["firm", "card", "date", "address", "item_name", "quantity", "price", "cost"]
    return await save_transactions(df, TransactionType.EXPENSE, document)

async def process_excel_payment(file_content: bytes, document: str):
    # New format for payments: Дата, карта, имя, вид транзакции, стоимость
    df = pd.read_excel(io.BytesIO(file_content), skiprows=2, usecols="B:F")
    df.columns = ["date", "card", "item_name", "type_str", "cost"]
    
    df["firm"] = ""
    df["address"] = ""
    df["quantity"] = 1.0
    df["price"] = df["cost"]
    
    return await save_transactions(df, TransactionType.PAYMENT, document)

async def save_transactions(df, t_type: TransactionType, document: str):
    """
    Сохраняем транзакции с логикой:
    - дубликат определяется по: карта, дата, тип, наименование, стоимость (округлённая до целого);
    - если в БД уже есть любая строка с той же картой и датой (даже если остальные поля отличаются),
      добавляем предупреждение с номером строки и 4 полями.
    """
    df = df.dropna(subset=["card", "date"])
    added_count = 0
    skipped_count = 0
    warnings = []

    async with async_session() as session:
        for idx, row in df.iterrows():
            card_number = str(row["card"])
            date = row["date"]

            if isinstance(date, str):
                for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
                    try:
                        date = datetime.strptime(date, fmt)
                        break
                    except Exception:
                        continue

            cost_val = float(row["cost"])
            rounded_cost = int(round(cost_val))
            item_name = str(row["item_name"])

            # Все существующие сделки с такой же картой и датой
            result = await session.execute(
                select(Transaction).where(
                    and_(
                        Transaction.card_number == card_number,
                        Transaction.date == date,
                    )
                )
            )
            existing = result.scalars().all()

            is_duplicate = False
            if existing:
                # Проверяем дубликат по всем 4 полям
                for t in existing:
                    if (
                        t.type == t_type
                        and (t.item_name or "") == item_name
                        and int(round(t.cost)) == rounded_cost
                    ):
                        is_duplicate = True
                        break

                # Всегда добавляем предупреждение о совпадении карты и даты
                excel_row = idx + 4  # данные начинаются с 4-й строки в Excel
                warnings.append(
                    {
                        "row": excel_row,
                        "card": card_number,
                        "date": date,
                        "item_name": item_name,
                        "cost_rounded": rounded_cost,
                    }
                )

            if is_duplicate:
                skipped_count += 1
                continue

            await add_to_whitelist(card_number)

            session.add(
                Transaction(
                    card_number=card_number,
                    document=document,
                    firm=str(row.get("firm", "")),
                    date=date,
                    address=str(row.get("address", "")),
                    item_name=item_name,
                    quantity=float(row.get("quantity", 0)),
                    price=float(row.get("price", 0)),
                    cost=cost_val,
                    type=t_type,
                )
            )
            added_count += 1

        await session.commit()

    return added_count, skipped_count, warnings

@router.message(AdminState.waiting_for_expense_file, F.document)
async def handle_expense_file(message: Message, state: FSMContext, bot: Bot):
    file_id = message.document.file_id
    file_name = message.document.file_name or "report"
    file = await bot.get_file(file_id)
    file_content = await bot.download_file(file.file_path)
    content_bytes = file_content.read()
    # формируем метку документа: первые 10 символов имени + текущий момент в SQL-формате
    now = datetime.now()
    document_label = f"{file_name[:10]}_{now.strftime('%Y-%m-%d %H:%M:%S')}"
    
    await state.update_data(file_bytes=content_bytes, document=document_label)
    
    if not await validate_format(content_bytes):
        await message.answer("Точно ли таблица в нужном формате? (B2 и A3 должны быть пустыми, B3 — заполнено)", 
                           reply_markup=get_confirm_format_kb("expense"))
        await state.set_state(AdminState.confirm_format_expense)
    else:
        await do_process_expense(message, state, content_bytes, document_label)

@router.message(AdminState.waiting_for_payment_file, F.document)
async def handle_payment_file(message: Message, state: FSMContext, bot: Bot):
    file_id = message.document.file_id
    file_name = message.document.file_name or "report"
    file = await bot.get_file(file_id)
    file_content = await bot.download_file(file.file_path)
    content_bytes = file_content.read()
    now = datetime.now()
    document_label = f"{file_name[:10]}_{now.strftime('%Y-%m-%d %H:%M:%S')}"
    
    await state.update_data(file_bytes=content_bytes, document=document_label)
    
    if not await validate_format(content_bytes):
        await message.answer("Точно ли таблица в нужном формате? (B2 и A3 должны быть пустыми, B3 — заполнено)", 
                           reply_markup=get_confirm_format_kb("payment"))
        await state.set_state(AdminState.confirm_format_payment)
    else:
        await do_process_payment(message, state, content_bytes, document_label)

@router.callback_query(F.data.startswith("confirm_yes_"))
async def confirm_yes(callback: CallbackQuery, state: FSMContext):
    report_type = callback.data.split("_")[-1]
    data = await state.get_data()
    content_bytes = data.get("file_bytes")
    document_label = data.get("document", "unknown")
    
    if report_type == "expense":
        await do_process_expense(callback.message, state, content_bytes, document_label)
    else:
        await do_process_payment(callback.message, state, content_bytes, document_label)
    await callback.answer()

@router.callback_query(F.data == "confirm_no")
async def confirm_no(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("Загрузка отменена. Пожалуйста, проверьте формат файла.")
    await state.clear()
    await callback.answer()

from bot.utils import update_last_update_time

async def do_process_expense(message, state, content_bytes, document_label: str):
    try:
        added, skipped, warnings = await process_excel_expense(content_bytes, document_label)
        update_last_update_time() # Обновляем время
        text = f"Обработка трат завершена.\nДобавлено: {added}\nПропущено (дубликаты): {skipped}"

        if warnings:
            text += "\n\n⚠️ Найдены строки с совпадающими номером карты и датой:\n"
            for w in warnings:
                dt_str = (
                    w["date"].strftime("%d.%m.%Y %H:%M")
                    if isinstance(w["date"], datetime)
                    else str(w["date"])
                )
                text += (
                    f"Строка {w['row']}: карта {w['card']}, дата {dt_str}, "
                    f"наименование/вид: {w['item_name']}, стоимость (округлённо): {w['cost_rounded']}\n"
                )

        await message.answer(text)
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
    await state.clear()

async def do_process_payment(message, state, content_bytes, document_label: str):
    try:
        added, skipped, warnings = await process_excel_payment(content_bytes, document_label)
        update_last_update_time() # Обновляем время
        text = f"Обработка оплат завершена.\nДобавлено: {added}\nПропущено (дубликаты): {skipped}"

        if warnings:
            text += "\n\n⚠️ Найдены строки с совпадающими номером карты и датой:\n"
            for w in warnings:
                dt_str = (
                    w["date"].strftime("%d.%m.%Y %H:%M")
                    if isinstance(w["date"], datetime)
                    else str(w["date"])
                )
                text += (
                    f"Строка {w['row']}: карта {w['card']}, дата {dt_str}, "
                    f"наименование/вид: {w['item_name']}, стоимость (округлённо): {w['cost_rounded']}\n"
                )

        await message.answer(text)
    except Exception as e:
        await message.answer(f"Ошибка: {e}")
    await state.clear()

@router.callback_query(F.data == "admin_export")
async def process_export_transactions(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return
    await callback.message.answer("Введите дату начала (ДД.ММ.ГГГГ):")
    await state.set_state(AdminState.waiting_for_export_start_date)
    await callback.answer()

@router.callback_query(F.data == "admin_broadcast")
async def broadcast_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return
    await callback.message.answer("Введите сообщение для рассылки всем пользователям:")
    await state.set_state(AdminState.waiting_for_broadcast_msg)
    await callback.answer()

@router.message(AdminState.waiting_for_broadcast_msg)
async def broadcast_process(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        return
    
    msg_text = message.text
    async with async_session() as session:
        from database.models import User
        result = await session.execute(select(User.telegram_id))
        user_ids = result.scalars().all()
    
    count = 0
    for uid in user_ids:
        try:
            await bot.send_message(uid, msg_text)
            count += 1
        except Exception:
            pass
    
    await message.answer(f"Рассылка завершена. Сообщение получили {count} пользователей.")
    await state.clear()

@router.callback_query(F.data == "admin_gen_link")
async def gen_link_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("Доступ запрещен")
        return
    await callback.message.answer("Введите номер карты, для которой нужно создать ссылку:")
    await state.set_state(AdminState.waiting_for_link_card)
    await callback.answer()

@router.message(AdminState.waiting_for_link_card)
async def gen_link_process(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        return
    
    card_number = message.text.strip()
    # Получаем имя бота для формирования ссылки
    bot_info = await bot.get_me()
    link = f"https://t.me/{bot_info.username}?start={card_number}"
    
    await message.answer(f"Ссылка для регистрации по карте {card_number}:\n\n<code>{link}</code>", parse_mode="HTML")
    await state.clear()

@router.message(AdminState.waiting_for_export_start_date)
async def process_start_date(message: Message, state: FSMContext):
    try:
        date = datetime.strptime(message.text.strip(), "%d.%m.%Y")
        await state.update_data(start_date=date)
        await message.answer("Введите дату окончания (ДД.ММ.ГГГГ):")
        await state.set_state(AdminState.waiting_for_export_end_date)
    except ValueError:
        await message.answer("Неверный формат даты. Используйте ДД.ММ.ГГГГ")

@router.message(AdminState.waiting_for_export_end_date)
async def process_end_date(message: Message, state: FSMContext):
    try:
        end_date = datetime.strptime(message.text.strip(), "%d.%m.%Y").replace(hour=23, minute=59, second=59)
        data = await state.get_data()
        start_date = data["start_date"]
        
        async with async_session() as session:
            result = await session.execute(
                select(Transaction).where(
                    and_(
                        Transaction.date >= start_date,
                        Transaction.date <= end_date
                    )
                ).order_by(Transaction.date)
            )
            transactions = result.scalars().all()
            
            if not transactions:
                await message.answer("За указанный период сделок не найдено.")
                await state.clear()
                return
            
            # Create Excel file
            df_data = []
            for t in transactions:
                df_data.append({
                    "Фирма": t.firm,
                    "Карта": t.card_number,
                    "Дата": t.date,
                    "Адрес": t.address,
                    "Наименование": t.item_name,
                    "Количество": t.quantity,
                    "Цена": t.price,
                    "Стоимость": t.cost,
                    "Тип": "Трата" if t.type == TransactionType.EXPENSE else "Оплата"
                })
            
            df = pd.DataFrame(df_data)
            output_file = f"export_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx"
            df.to_excel(output_file, index=False)
            
            await message.answer_document(FSInputFile(output_file))
            os.remove(output_file)
            
        await state.clear()
    except ValueError:
        await message.answer("Неверный формат даты. Используйте ДД.ММ.ГГГГ")
    except Exception as e:
        await message.answer(f"Ошибка при выгрузке: {e}")
        await state.clear()
